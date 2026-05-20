"""
Generate realistic DWA V&V test data spreadsheet for demo use.
Run: python3 docs/generate_test_data.py
Output: docs/DWA_VV_Test_Data.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Palette (DWS brand colours) ──────────────────────────────────────────────
BLUE   = "004A97"   # DWS navy
CYAN   = "00AEEF"   # DWS light blue
GREEN  = "00843D"   # DWS green
WHITE  = "FFFFFF"
LGREY  = "F2F2F2"
MGREY  = "D9D9D9"
ORANGE = "E87722"
RED    = "C8102E"

def hdr(text, bold=True, size=11, color=WHITE, bg=BLUE):
    return {"value": text, "font": Font(bold=bold, size=size, color=color),
            "fill": PatternFill("solid", fgColor=bg),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)}

def cell(text, bold=False, bg=None, color="000000", align="left", wrap=False):
    c = {"value": text,
         "font": Font(bold=bold, color=color),
         "alignment": Alignment(horizontal=align, vertical="center", wrap_text=wrap)}
    if bg:
        c["fill"] = PatternFill("solid", fgColor=bg)
    return c

def status_cell(text):
    colour_map = {
        "Not Commenced": MGREY,
        "In Process": CYAN,
        "Completed": GREEN,
        "CP1 – Inception": ORANGE,
        "CP2 – Spatial Info": ORANGE,
        "CP3 – WARMS Eval": ORANGE,
        "CP4 – Add. Info": ORANGE,
        "CP5 – GIS / Mapbooks": CYAN,
        "CP6 – SAPWAT": CYAN,
        "CP7 – ELU Calc": "8B0000",
        "CP8 – Dam Volumes": "8B0000",
        "CP9 – SFRA Calc": "8B0000",
        "Letter 1 Issued": ORANGE,
        "ELU Confirmed": GREEN,
        "S33(2) Declared": GREEN,
    }
    bg = colour_map.get(text, LGREY)
    txt_col = WHITE if bg not in (MGREY, LGREY, CYAN) else "000000"
    return cell(text, bold=True, bg=bg, color=txt_col, align="center")

def write_sheet(ws, title_row, headers, rows, col_widths):
    """Write a title row, then header row, then data rows."""
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tc = ws.cell(row=1, column=1)
    tc.value = title_row
    tc.font = Font(bold=True, size=13, color=WHITE)
    tc.fill = PatternFill("solid", fgColor=BLUE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col)
        for k, v in hdr(h).items():
            setattr(c, k, v)
        c.border = border
    ws.row_dimensions[2].height = 36

    # Data rows
    for r, row in enumerate(rows, 3):
        bg = LGREY if r % 2 == 0 else WHITE
        for col, cell_def in enumerate(row, 1):
            c = ws.cell(row=r, column=col)
            if isinstance(cell_def, dict):
                for k, v in cell_def.items():
                    setattr(c, k, v)
                if not cell_def.get("fill"):
                    c.fill = PatternFill("solid", fgColor=bg)
            else:
                c.value = cell_def
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(vertical="center")
            c.border = border
        ws.row_dimensions[r].height = 20

    # Column widths
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════════════
# DATA
# ═══════════════════════════════════════════════════════════════════════════════

PROPERTIES = [
    # SGCode,          Farm Name,                  Province,    WMA,               Catchment, Lat,       Lon,        Status
    ("LP 45 KR",       "Rooipoort 45",             "Limpopo",   "Olifants",        "B72A",    "-24.1823", "30.4156", "Active"),
    ("LP 234 LQ",      "Welgelegen 234",           "Limpopo",   "Olifants",        "B71B",    "-24.5601", "29.8834", "Active"),
    ("MP 89 MS",       "Cyferfontein 89",          "Mpumalanga","Inkomati-Usuthu",  "X31A",    "-25.3127", "31.2045", "Active"),
    ("MP 456 NR",      "Hartbeeshoek 456",         "Mpumalanga","Inkomati-Usuthu",  "X32C",    "-25.0884", "31.5672", "Active"),
    ("MP 12 KP",       "Klipplaatdrift 12",        "Mpumalanga","Inkomati-Usuthu",  "X31B",    "-25.7219", "30.9341", "Active"),
    ("GP 321 IQ",      "Modderfontein 321",        "Gauteng",   "Upper Vaal",      "C11A",    "-26.4103", "28.4557", "Active"),
    ("NW 67 HO",       "De Hoop 67",               "North West","Upper Vaal",      "C12B",    "-26.8901", "27.1234", "Active"),
    ("KZN 789 HV",     "Springfontein 789",        "KwaZulu-Natal","Thukela",      "V20A",    "-28.4512", "29.6723", "Active"),
    ("KZN 101 BN",     "Tweefontein 101",          "KwaZulu-Natal","Thukela",      "V21B",    "-28.9034", "30.1890", "Active"),
    ("LP 555 PT",      "Boomplaats 555",           "Limpopo",   "Limpopo",         "A92A",    "-23.4501", "29.2378", "Active"),
    ("LP 203 WR",      "Goedehoop 203",            "Limpopo",   "Limpopo",         "A91B",    "-23.7812", "28.8934", "Consolidated"),
    ("MP 77 BB",       "Vlakfontein 77",           "Mpumalanga","Olifants",        "B73C",    "-24.9923", "30.6711", "Active"),
]

OWNERS = [
    # Owner No, Title, First, Last,       ID Number,        Entity Type,   Farm (SG),    Title Deed No,     Date
    ("OWN-001", "Mr",  "Johannes",  "van der Merwe",  "5803145678089", "Individual",  "LP 45 KR",   "T12345/2001",     "2001-03-15"),
    ("OWN-002", "Ms",  "Thandi",    "Nkosi",          "7206220987654", "Individual",  "LP 234 LQ",  "T67890/2003",     "2003-07-22"),
    ("OWN-003", "Mr",  "Pieter",    "Boshoff",        "6501085432109", "Individual",  "MP 89 MS",   "T24680/1998",     "1998-11-08"),
    ("OWN-004", "",    "Hartbees",  "Boerdery (Pty) Ltd", "N/A",        "Company",     "MP 456 NR",  "T11223/2005",     "2005-02-14"),
    ("OWN-005", "Mrs", "Maria",     "Dlamini",        "7512145678901", "Individual",  "MP 12 KP",   "T99001/2000",     "2000-09-30"),
    ("OWN-006", "Mr",  "Hendrik",   "Groenewald",     "5911285432101", "Individual",  "GP 321 IQ",  "T55443/1996",     "1996-04-17"),
    ("OWN-007", "",    "De Hoop",   "Landgoed BK",    "N/A",           "CC",          "NW 67 HO",   "T33221/2010",     "2010-06-05"),
    ("OWN-008", "Mr",  "Sipho",     "Mthembu",        "6804175678091", "Individual",  "KZN 789 HV", "T77665/1999",     "1999-01-20"),
    ("OWN-009", "Ms",  "Lerato",    "Molefe",         "8203145678012", "Individual",  "KZN 101 BN", "T44332/2007",     "2007-08-11"),
    ("OWN-010", "Mr",  "Gert",      "Swanepoel",      "5507095432108", "Individual",  "LP 555 PT",  "T22110/1993",     "1993-12-01"),
    ("OWN-011", "",    "Goedehoop", "Plaas (Pty) Ltd","N/A",           "Company",     "LP 203 WR",  "T88776/2002",     "2002-05-28"),
    ("OWN-012", "Mr",  "Andries",   "Pretorius",      "6309185432107", "Individual",  "MP 77 BB",   "T66554/1997",     "1997-10-04"),
]

CASES = [
    # Case No,        SGCode,     WARMS Reg No,   Validation Status,       Workflow State,              Assign To,    WMA,              GWCA?,  Track,     Date Created
    ("VV-2024-0001",  "LP 45 KR",  "W/MP/B72A/45/001", "In Process",      "CP7 – ELU Calc",            "Jane Validator","Olifants",    "No",   "S35",     "2024-03-01"),
    ("VV-2024-0002",  "LP 234 LQ", "W/MP/B71B/234/002","In Process",      "CP5 – GIS / Mapbooks",      "Jane Validator","Olifants",    "No",   "S35",     "2024-03-15"),
    ("VV-2024-0003",  "MP 89 MS",  "W/MP/X31A/89/003", "In Process",      "CP6 – SAPWAT",              "Jane Validator","Inkomati-Usuthu","No","S35",     "2024-04-02"),
    ("VV-2024-0004",  "MP 456 NR", "W/MP/X32C/456/004","Completed",       "ELU Confirmed",             "Jane Validator","Inkomati-Usuthu","No","S35",     "2024-01-10"),
    ("VV-2024-0005",  "MP 12 KP",  "W/MP/X31B/12/005", "Not Commenced",   "CP1 – Inception",           "Jane Validator","Inkomati-Usuthu","No","S35",     "2024-05-20"),
    ("VV-2024-0006",  "GP 321 IQ", "W/GP/C11A/321/006","In Process",      "CP3 – WARMS Eval",          "Jane Validator","Upper Vaal",   "No",   "S35",     "2024-02-28"),
    ("VV-2024-0007",  "NW 67 HO",  "W/NW/C12B/67/007", "In Process",      "CP4 – Add. Info",           "Jane Validator","Upper Vaal",   "No",   "S35",     "2024-03-22"),
    ("VV-2024-0008",  "KZN 789 HV","W/KZN/V20A/789/008","In Process",     "CP8 – Dam Volumes",         "Jane Validator","Thukela",      "No",   "S35",     "2024-02-01"),
    ("VV-2024-0009",  "KZN 101 BN","W/KZN/V21B/101/009","In Process",     "Letter 1 Issued",           "Jane Validator","Thukela",      "No",   "S35",     "2023-11-14"),
    ("VV-2024-0010",  "LP 555 PT", "W/LP/A92A/555/010", "In Process",     "CP2 – Spatial Info",        "Jane Validator","Limpopo",      "No",   "S35",     "2024-06-01"),
    ("VV-2024-0011",  "LP 203 WR", "W/LP/A91B/203/011", "Completed",      "S33(2) Declared",           "Jane Validator","Limpopo",      "No",   "S33(2)",  "2023-09-05"),
    ("VV-2024-0012",  "MP 77 BB",  "W/MP/B73C/77/012",  "In Process",     "CP9 – SFRA Calc",           "Jane Validator","Olifants",     "No",   "S35",     "2024-01-30"),
]

IRRIGATION = [
    # Case No,         Field No, Crop,              Area (ha), Irr System,           Water Source,    Volume (m³/a), SAPWAT (mm/ha/a), QP Exists?
    ("VV-2024-0001",   1,  "Maize",                  18.5,  "Centre Pivot",        "River",          185000,          8200,           "Yes"),
    ("VV-2024-0001",   2,  "Wheat",                   9.2,  "Flood Irrigation",    "River",           73600,          7400,           "Yes"),
    ("VV-2024-0002",   1,  "Tomatoes",               12.0,  "Drip Irrigation",     "Borehole",        96000,          7900,           "Yes"),
    ("VV-2024-0002",   2,  "Potatoes",                8.5,  "Sprinkler",           "River",           68000,          8500,           "Yes"),
    ("VV-2024-0003",   1,  "Sugarcane",              35.0,  "Flood Irrigation",    "River",          595000,         14200,           "Yes"),
    ("VV-2024-0003",   2,  "Lucerne",                 6.0,  "Sprinkler",           "River",           72000,         12000,           "Yes"),
    ("VV-2024-0004",   1,  "Citrus",                 22.0,  "Micro-Sprinkler",     "River",          220000,          9200,           "Yes"),
    ("VV-2024-0004",   2,  "Avocado",                10.0,  "Drip Irrigation",     "Dam",             80000,          7800,           "Yes"),
    ("VV-2024-0006",   1,  "Maize",                  40.0,  "Centre Pivot",        "River",          400000,          8300,           "Yes"),
    ("VV-2024-0007",   1,  "Maize",                  25.0,  "Centre Pivot",        "River",          250000,          8100,           "No"),
    ("VV-2024-0007",   2,  "Soya Beans",             15.0,  "Flood Irrigation",    "River",          105000,          7200,           "No"),
    ("VV-2024-0009",   1,  "Pasture",                50.0,  "Flood Irrigation",    "River",          500000,         10000,           "Yes"),
    ("VV-2024-0010",   1,  "Maize",                  30.0,  "Centre Pivot",        "River",          300000,          8400,           "Yes"),
    ("VV-2024-0012",   1,  "Eucalyptus",             45.0,  "N/A (SFRA)",          "N/A",                 0,              0,           "Yes"),
    ("VV-2024-0012",   2,  "Pine",                   28.0,  "N/A (SFRA)",          "N/A",                 0,              0,           "Yes"),
]

DAM_CALC = [
    # Case No,          Dam Name,          Method, Wall Len(m), Fetch(m), R1(m),  C1(m), Factor, Area(ha), Depth(m), Capacity(m³)
    ("VV-2024-0004",   "Hartbees Dam 1",   "1 (Wall Length)", 120,  85,   1200,   15,   0.40,  None,    None,   204000),
    ("VV-2024-0008",   "Springfontein Dam","1 (Wall Length)", 95,   60,    900,   12,   0.33,  None,    None,    94050),
    ("VV-2024-0008",   "Springer Dam 2",   "2 (Area)",        None, None,  None,  None,  0.50,  3.5,     2.8,     49000),
    ("VV-2024-0009",   "Tweefontein Dam",  "1 (Wall Length)", 150,  100,  1500,   18,   0.40,  None,    None,   333333),
]

FORESTATION = [
    # Case No,           Species,      Pre1972?(ha), Post1984 Permit ha, QP ha,  Current ha, Lawful Vol(m³/a), Unlawful Vol
    ("VV-2024-0012",    "Eucalyptus",  0,            0,                  45.0,   45.0,        90000,           0),
    ("VV-2024-0012",    "Pine",        28.0,         0,                  28.0,   28.0,        56000,           0),
]


# ═══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

wb = Workbook()

# ── Sheet 1: Properties ──────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Properties"
prop_headers = [
    "SG Code", "Farm Name", "Province", "Water Management Area",
    "Quaternary Catchment", "Latitude", "Longitude",
    "Property Status", "Notes"
]
prop_rows = []
for p in PROPERTIES:
    sg, name, prov, wma, cat, lat, lon, status = p
    note = "Consolidated — parent record retained" if status == "Consolidated" else ""
    sc = status_cell(status) if status == "Consolidated" else cell(status, bg=LGREY if status == "Active" else ORANGE)
    prop_rows.append([
        cell(sg, bold=True), cell(name), cell(prov), cell(wma),
        cell(cat, align="center"), cell(lat, align="right"),
        cell(lon, align="right"), sc, cell(note, color="666666")
    ])
write_sheet(ws1, "DWA V&V — Property Register", prop_headers, prop_rows,
            [14, 22, 16, 22, 16, 12, 12, 16, 34])

# ── Sheet 2: Property Owners ─────────────────────────────────────────────────
ws2 = wb.create_sheet("Property Owners")
own_headers = [
    "Owner No", "Title", "First Name", "Last Name / Entity Name",
    "ID / Reg Number", "Entity Type", "Linked Farm (SG Code)",
    "Title Deed No", "Registration Date"
]
own_rows = []
for o in OWNERS:
    no, title, first, last, idno, etype, sg, deed, date = o
    own_rows.append([
        cell(no, bold=True), cell(title), cell(first), cell(last, bold=True),
        cell(idno), cell(etype),
        cell(sg, bold=True, color=BLUE), cell(deed), cell(date)
    ])
write_sheet(ws2, "DWA V&V — Property Owner Register", own_headers, own_rows,
            [12, 8, 14, 26, 18, 14, 18, 16, 16])

# ── Sheet 3: V&V Cases ───────────────────────────────────────────────────────
ws3 = wb.create_sheet("V&V Cases")
case_headers = [
    "Case No", "SG Code", "WARMS Reg No", "Validation Status",
    "Current Workflow Stage", "Assigned Validator",
    "WMA", "GWCA?", "Assessment Track", "Date Created"
]
case_rows = []
for c in CASES:
    cno, sg, warms, vstatus, wstage, asn, wma, gwca, track, dt = c
    case_rows.append([
        cell(cno, bold=True, color=BLUE),
        cell(sg, bold=True),
        cell(warms),
        status_cell(vstatus),
        status_cell(wstage),
        cell(asn),
        cell(wma),
        cell(gwca, align="center", bold=(gwca == "Yes")),
        cell(track, bold=True),
        cell(dt)
    ])
write_sheet(ws3, "DWA V&V — Case Register (FileMaster)", case_headers, case_rows,
            [16, 14, 24, 18, 22, 18, 22, 8, 12, 14])

# ── Sheet 4: Irrigation & Crops ──────────────────────────────────────────────
ws4 = wb.create_sheet("Irrigation & Crops")
irr_headers = [
    "Case No", "Field No", "Crop Type", "Field Area (ha)",
    "Irrigation System", "Water Source",
    "Annual Volume (m³/a)", "SAPWAT Result (mm/ha/a)",
    "Qualifying Period Evidence?"
]
irr_rows = []
for i in IRRIGATION:
    cno, fno, crop, area, sys_, src, vol, sapwat, qp = i
    qp_cell = cell("YES", bold=True, color=GREEN, align="center") if qp == "Yes" \
              else cell("NO", bold=True, color=RED, align="center")
    irr_rows.append([
        cell(cno, bold=True, color=BLUE),
        cell(fno, align="center"),
        cell(crop),
        cell(area, align="right"),
        cell(sys_),
        cell(src),
        cell(f"{vol:,}", align="right"),
        cell(sapwat if sapwat else "N/A", align="right"),
        qp_cell
    ])
write_sheet(ws4, "DWA V&V — Field & Crop / Irrigation Data", irr_headers, irr_rows,
            [16, 10, 16, 16, 22, 16, 22, 22, 24])

# ── Sheet 5: Dam Calculations ────────────────────────────────────────────────
ws5 = wb.create_sheet("Dam Calculations")
dam_headers = [
    "Case No", "Dam Name", "Calc Method",
    "Wall Length (m)", "Fetch (m)", "River Dist R1 (m)", "Contour Diff C1 (m)",
    "Shape Factor", "Area (ha)", "Depth (m)", "Capacity (m³)"
]
dam_rows = []
for d in DAM_CALC:
    cno, name, method, wl, fetch, r1, c1, factor, area, depth, cap = d
    dam_rows.append([
        cell(cno, bold=True, color=BLUE), cell(name, bold=True),
        cell(method),
        cell(wl if wl else "—", align="right"),
        cell(fetch if fetch else "—", align="right"),
        cell(r1 if r1 else "—", align="right"),
        cell(c1 if c1 else "—", align="right"),
        cell(factor, align="center"),
        cell(area if area else "—", align="right"),
        cell(depth if depth else "—", align="right"),
        cell(f"{cap:,}", bold=True, align="right", color=BLUE),
    ])
write_sheet(ws5, "DWA V&V — Dam Volume Calculations (Appendix D)", dam_headers, dam_rows,
            [16, 22, 20, 16, 12, 18, 18, 14, 12, 12, 16])

# ── Sheet 6: Forestation / SFRA ──────────────────────────────────────────────
ws6 = wb.create_sheet("Forestation (SFRA)")
sfra_headers = [
    "Case No", "Species", "Pre-1972 Area (ha)",
    "Post-1984 Permit Area (ha)", "Qualifying Period Area (ha)",
    "Current Area (ha)", "Lawful Vol (m³/a)", "Unlawful Vol (m³/a)"
]
sfra_rows = []
for f in FORESTATION:
    cno, sp, pre72, post84, qpha, curha, lawful, unlawful = f
    sfra_rows.append([
        cell(cno, bold=True, color=BLUE), cell(sp, bold=True),
        cell(pre72, align="right"),
        cell(post84 if post84 else "—", align="right"),
        cell(qpha, align="right"),
        cell(curha, align="right"),
        cell(f"{lawful:,}", align="right", color=GREEN, bold=True),
        cell(f"{unlawful:,}" if unlawful else "—", align="right",
             color=RED if unlawful else "000000"),
    ])
write_sheet(ws6, "DWA V&V — Stream Flow Reduction Activities (SFRA / S21d)", sfra_headers, sfra_rows,
            [16, 16, 20, 24, 26, 18, 20, 20])

# ── Sheet 7: File Checklist ───────────────────────────────────────────────────
ws7 = wb.create_sheet("File Checklist")
chk_headers = [
    "Case No",
    "1. Checklist", "2. WARMS Report", "3. Title Deed Report",
    "4. SG Diagram", "5. Previous Study/Leg Docs",
    "6. Field & Crop Summary", "7. QP Mapbook",
    "8. QP Volume Calcs", "9. Current Mapbook",
    "10. Current Vol Calcs", "11. Water Storage Calcs",
    "12. S35/S33 Letters"
]

def tick(val):
    if val == "Y":  return cell("✓", bold=True, color=GREEN, align="center", bg="E8F5E9")
    if val == "N":  return cell("✗", bold=True, color=RED, align="center", bg="FFEBEE")
    return cell("—", align="center", color="AAAAAA")

checklist_data = [
    #                         1    2    3    4    5    6    7    8    9    10   11   12
    ("VV-2024-0001", "Y","Y","Y","Y","Y","Y","Y","Y","Y","N","N","N"),
    ("VV-2024-0002", "Y","Y","Y","Y","Y","Y","Y","N","N","N","N","N"),
    ("VV-2024-0003", "Y","Y","Y","Y","Y","Y","Y","Y","Y","N","N","N"),
    ("VV-2024-0004", "Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y"),
    ("VV-2024-0005", "Y","Y","N","N","N","N","N","N","N","N","N","N"),
    ("VV-2024-0006", "Y","Y","Y","N","N","N","N","N","N","N","N","N"),
    ("VV-2024-0007", "Y","Y","Y","Y","Y","N","N","N","N","N","N","N"),
    ("VV-2024-0008", "Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","N"),
    ("VV-2024-0009", "Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y"),
    ("VV-2024-0010", "Y","Y","Y","N","N","N","N","N","N","N","N","N"),
    ("VV-2024-0011", "Y","Y","Y","Y","Y","N","N","N","N","N","N","Y"),
    ("VV-2024-0012", "Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","N","N"),
]
chk_rows = []
for row in checklist_data:
    cno = row[0]
    chk_rows.append([cell(cno, bold=True, color=BLUE)] + [tick(v) for v in row[1:]])
write_sheet(ws7, "DWA V&V — File Completeness Checklist (Appendix A)", chk_headers, chk_rows,
            [16] + [14]*12)

# ── Cover sheet ───────────────────────────────────────────────────────────────
ws0 = wb.create_sheet("README", 0)
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 2
ws0.column_dimensions["B"].width = 60
ws0.column_dimensions["C"].width = 40

rows_info = [
    (1,  "DWA Validation & Verification System", True, 16, WHITE, BLUE, 36),
    (2,  "Test Data Workbook — Demo Use Only",    True, 12, WHITE, BLUE, 24),
    (3,  "",                                      False, 11, "000000", WHITE, 8),
    (4,  "Purpose",                               True, 12, WHITE, CYAN, 22),
    (5,  "Realistic test data for DWA V&V system demo. 12 properties across", False, 11, "000000", LGREY, 18),
    (6,  "Olifants, Inkomati-Usuthu, Upper Vaal, Thukela & Limpopo WMAs.",   False, 11, "000000", LGREY, 18),
    (7,  "",                                      False, 11, "000000", WHITE, 8),
    (8,  "Sheets",                                True, 12, WHITE, CYAN, 22),
    (9,  "1. Properties       — 12 farm parcels with SG codes & coordinates",  False, 11, "000000", LGREY, 18),
    (10, "2. Property Owners  — 12 owners (individuals & entities)",            False, 11, "000000", WHITE, 18),
    (11, "3. V&V Cases        — 12 FileMaster cases at different workflow stages",False,11,"000000",LGREY, 18),
    (12, "4. Irrigation & Crops — 15 field/crop records with SAPWAT results",   False, 11, "000000", WHITE, 18),
    (13, "5. Dam Calculations — 4 dam records (Appendix D methods)",            False, 11, "000000", LGREY, 18),
    (14, "6. Forestation (SFRA) — S21(d) stream flow records",                  False, 11, "000000", WHITE, 18),
    (15, "7. File Checklist   — Appendix A completeness per case",              False, 11, "000000", LGREY, 18),
    (16, "",                                      False, 11, "000000", WHITE, 8),
    (17, "Login Credentials (Azure Demo)",        True, 12, WHITE, GREEN, 22),
    (18, "URL:      https://dwa-vv-demo.azurewebsites.net", False, 11, "000000", LGREY, 18),
    (19, "Admin:    admin@dwa.demo        / Demo@Pass2026", False, 11, "000000", WHITE, 18),
    (20, "Validator: validator-{wma}@dwa.demo  / Demo@Pass2026", False, 11, "000000", LGREY, 18),
    (21, "Capturer: capturer-{wma}@dwa.demo  / Demo@Pass2026",  False, 11, "000000", WHITE, 18),
    (22, "",                                      False, 11, "000000", WHITE, 8),
    (23, "Qualifying Period: 1 October 1996 – 30 September 1998", False, 11, "666666", LGREY, 18),
    (24, "Generated: 2026-05-07  |  For demo purposes only — not real water use data", False, 10, "999999", WHITE, 16),
]
ws0.merge_cells("B1:C1")
ws0.merge_cells("B2:C2")
for row_num, text, bold, size, fg, bg, height in rows_info:
    ws0.merge_cells(f"B{row_num}:C{row_num}")
    c = ws0.cell(row=row_num, column=2)
    c.value = text
    c.font = Font(bold=bold, size=size, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws0.row_dimensions[row_num].height = height

out = "docs/DWA_VV_Test_Data.xlsx"
wb.save(out)
print(f"Saved: {out}")
